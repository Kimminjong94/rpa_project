from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 영어 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]


#a열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5


#스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) #글자 색은 빨갛게, 이탤릭, 볼드체
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single") #글자크기 20으로 밑줄적용


#테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        #center, left, right, top, bottom

        if cell.column == 1: #A번호열은 제외
            continue

        #cell이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") #배경을 초록색으로 설정
            cell.font = Font(color="FF0000")

#틀고정
ws.freeze_panes = "B2" #B2기준 틀고정

wb.save("sample_style.xlsx")