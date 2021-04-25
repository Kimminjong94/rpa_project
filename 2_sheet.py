from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" #시트이름 변경
ws.sheet_properties.tabColor = "9900ff" #rgb형태로 색 바꿔줌


#sheet mysheet yoursheet
ws1 = wb.create_sheet("yoursheet") #주어진 이름으로 시트 생성
ws2 = wb.create_sheet("newsheet", 2) #2번째 index에 sheet 생성


new_ws = wb["newsheet"] #dictionary 형태로 sheet 에 접근가능 

print(wb.sheetnames) #모든 시트 이름 확인

#sheet 복사 
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")